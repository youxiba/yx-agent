# coding=utf-8
import csv
import io
from openpyxl import load_workbook
from ..file import File
from ..spi import BaseSplitHandle, ParagraphRaw, SplitOptions


class TableSplitHandle(BaseSplitHandle):
    """表格导入：每行转一条段落，首列作为标题，形如「列名: 值」拼接内容"""

    def support(self, file: File) -> bool:
        return file.meta.get("source_type") == "table" and file.suffix in {".xlsx", ".xls", ".csv"}

    def get_content(self, file: File, save_image: bool = False) -> str:
        return ""

    def handle(self, file: File, opts: SplitOptions) -> list[ParagraphRaw]:
        if file.suffix == ".csv":
            text = file.open_bytes().decode("utf-8-sig", errors="ignore")
            rows = list(csv.reader(io.StringIO(text)))
        else:
            wb = load_workbook(file.open_workbook_source(), read_only=True)
            rows = [r for ws in wb.worksheets for r in ws.iter_rows(values_only=True)]
        header = [str(c) if c is not None else "" for c in rows[0]] if rows else []
        raws = []
        for r in rows[1:]:
            cells = [str(c) if c is not None else "" for c in r]
            content = " | ".join(f"{h}: {v}" for h, v in zip(header, cells) if v)
            raws.append(ParagraphRaw(title=cells[0] if cells else "", content=content))
        return raws