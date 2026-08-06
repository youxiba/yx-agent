# coding=utf-8
import csv
import io
from openpyxl import load_workbook
from ..file import File
from ..spi import BaseSplitHandle, ParagraphRaw, SplitOptions


class QaSplitHandle(BaseSplitHandle):
    """QA 导入：md 每段（空行分隔）= 一条问答（首行问题、余下回答）；xlsx/csv 两列「问题, 回答」"""

    def support(self, file: File) -> bool:
        return file.meta.get("source_type") == "qa" and file.suffix in {".md", ".txt", ".xlsx", ".xls", ".csv"}

    def get_content(self, file: File, save_image: bool = False) -> str:
        return ""

    def handle(self, file: File, opts: SplitOptions) -> list[ParagraphRaw]:
        if file.suffix in {".md", ".txt"}:
            pairs = self._parse_md(file.open_bytes().decode("utf-8-sig", errors="ignore"))
        else:
            pairs = self._parse_spreadsheet(file)
        return [ParagraphRaw(title=q[:256], content=a, questions=[q]) for q, a in pairs if q]

    @staticmethod
    def _parse_md(text: str) -> list[tuple[str, str]]:
        pairs = []
        for block in text.split("\n\n"):
            lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
            if not lines:
                continue
            pairs.append((lines[0], "\n".join(lines[1:])))
        return pairs

    @staticmethod
    def _parse_spreadsheet(file: File) -> list[tuple[str, str]]:
        if file.suffix == ".csv":
            text = file.open_bytes().decode("utf-8-sig", errors="ignore")
            rows = list(csv.reader(io.StringIO(text)))
        else:
            wb = load_workbook(file.path, read_only=True)
            rows = [r for ws in wb.worksheets for r in ws.iter_rows(values_only=True)]
        pairs = []
        for r in rows[1:]:          # 跳过表头
            if r and r[0] is not None:
                pairs.append((str(r[0]).strip(), str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""))
        return pairs