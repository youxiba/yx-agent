# coding=utf-8
from openpyxl import load_workbook
from ..file import File
from ..spi import BaseSplitHandle, ParagraphRaw, SplitOptions
from ..split_model import SplitModel


class XlsxSplitHandle(BaseSplitHandle):
    def support(self, file: File) -> bool:
        return file.suffix in {".xlsx", ".xls"} and file.meta.get("source_type") != "table"

    def get_content(self, file: File, save_image: bool = False) -> str:
        wb = load_workbook(file.open_workbook_source(), read_only=True)
        parts = []
        for ws in wb.worksheets:
            parts.append(f"# {ws.title}\n{self._sheet_to_md(ws)}")
        return "\n\n".join(parts)

    def handle(self, file: File, opts: SplitOptions) -> list[ParagraphRaw]:
        wb = load_workbook(file.open_workbook_source(), read_only=True)
        sm = SplitModel(opts.content_level_pattern, opts.limit, opts.with_filter)
        raws: list[ParagraphRaw] = []
        for ws in wb.worksheets:
            raws.extend(sm.parse_to_paragraphs(self._sheet_to_md(ws), title=ws.title))
        return raws

    @staticmethod
    def _sheet_to_md(ws) -> str:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ""
        header = [str(c) if c is not None else "" for c in rows[0]]
        md = "| " + " | ".join(header) + " |\n| " + " | ".join("---" for _ in header) + " |\n"
        for r in rows[1:]:
            md += "| " + " | ".join(str(c) if c is not None else "" for c in r) + " |\n"
        return md